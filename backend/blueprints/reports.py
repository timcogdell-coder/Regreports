import io
import calendar
from collections import defaultdict
from flask import Blueprint, request, jsonify, send_file
from flask_login import login_required
from models import db, Sample, SampleResult, PermitLimit, Parameter, Company, Permit, Violation, MonthlyFlowReport

reports_bp = Blueprint("reports", __name__)


@reports_bp.route("/samples", methods=["GET"])
@login_required
def sample_report():
    company_id   = request.args.get("company_id",   type=int)
    parameter_id = request.args.get("parameter_id", type=int)
    start_date   = request.args.get("start_date")
    end_date     = request.args.get("end_date")

    query = (
        db.session.query(SampleResult, Sample, PermitLimit, Parameter, Company, Permit)
        .join(Sample,      SampleResult.sample_id       == Sample.id)
        .join(PermitLimit, SampleResult.permit_limit_id == PermitLimit.id)
        .join(Parameter,   PermitLimit.parameter_id     == Parameter.id)
        .join(Company,     Sample.company_id             == Company.id)
        .join(Permit,      Sample.permit_id              == Permit.id)
    )

    if company_id:
        query = query.filter(Sample.company_id == company_id)
    if parameter_id:
        query = query.filter(PermitLimit.parameter_id == parameter_id)
    if start_date:
        query = query.filter(Sample.sample_date >= start_date)
    if end_date:
        query = query.filter(Sample.sample_date <= end_date)

    rows = query.order_by(
        Sample.sample_date,
        Company.name,
        Parameter.name,
    ).all()

    result = []
    for sr, s, pl, p, c, permit in rows:
        if pl.is_monitor_report:
            status = "MR"
        elif pl.is_range_limit:
            val = sr.concentration_result
            if val is None:
                status = "—"
            elif pl.min_value is not None and pl.max_value is not None:
                status = "Pass" if pl.min_value <= val <= pl.max_value else "Exceedance"
            else:
                status = "—"
        else:
            exceeds_conc = (pl.daily_max_concentration is not None and
                            sr.concentration_result is not None and
                            sr.concentration_result > pl.daily_max_concentration)
            exceeds_load = (pl.daily_max_loading is not None and
                            sr.loading_result is not None and
                            sr.loading_result > pl.daily_max_loading)
            status = "Exceedance" if (exceeds_conc or exceeds_load) else "Pass"

        result.append({
            "monitoring_period": s.sample_date.strftime("%B %Y"),
            "company_name":      c.name,
            "permit_number":     permit.permit_number,
            "sample_date":       str(s.sample_date),
            "sampler_name":      s.sampler_name,
            "flow_mgd":          s.flow_mgd,
            "parameter_name":    p.name,
            "abbreviation":      p.abbreviation,
            "concentration":     sr.concentration_result,
            "loading":           sr.loading_result,
            "limit_conc":        pl.daily_max_concentration,
            "limit_load":        pl.daily_max_loading,
            "is_monitor_report": pl.is_monitor_report or False,
            "is_range_limit":    pl.is_range_limit or False,
            "min_value":         pl.min_value,
            "max_value":         pl.max_value,
            "status":            status,
        })

    return jsonify(result), 200


@reports_bp.route("/monthly", methods=["GET"])
@login_required
def monthly_report():
    """
    Discharge Monitoring Report (DMR) — one row per permit limit for a company/month.
    Returns sample count, exceedance count, min/avg/max measured values, and limit info.
    """
    from sqlalchemy import func

    company_id = request.args.get("company_id", type=int)
    month      = request.args.get("month",      type=int)
    year       = request.args.get("year",       type=int)

    if not all([company_id, month, year]):
        return jsonify({"error": "company_id, month, and year are required"}), 400

    company = Company.query.get_or_404(company_id)

    # All active permits for the company
    permits = Permit.query.filter_by(company_id=company_id, is_active=True).all()
    if not permits:
        return jsonify({"company_name": company.name, "rows": []}), 200

    permit_ids = [p.id for p in permits]

    # Fetch every sample result in this month for this company
    raw = (
        db.session.query(SampleResult, Sample, PermitLimit)
        .join(Sample,      SampleResult.sample_id       == Sample.id)
        .join(PermitLimit, SampleResult.permit_limit_id == PermitLimit.id)
        .filter(
            Sample.company_id == company_id,
            func.extract("month", Sample.sample_date) == month,
            func.extract("year",  Sample.sample_date) == year,
        )
        .all()
    )

    # Group results by permit_limit_id
    groups: dict = defaultdict(list)
    for sr, s, pl in raw:
        groups[pl.id].append((sr, s))

    # All permit limits across all active permits (sorted by parameter name)
    all_limits = (
        PermitLimit.query
        .filter(PermitLimit.permit_id.in_(permit_ids))
        .join(Parameter, PermitLimit.parameter_id == Parameter.id)
        .order_by(Parameter.name)
        .all()
    )

    # Reviewed flow report for this month (authoritative source for flow limits)
    flow_report = MonthlyFlowReport.query.filter_by(
        company_id=company_id,
        report_month=month,
        report_year=year,
        review_status="reviewed",
    ).first()

    # Also check for any (pending/rejected) report so we can show its status
    any_flow_report = flow_report or MonthlyFlowReport.query.filter_by(
        company_id=company_id,
        report_month=month,
        report_year=year,
    ).first()

    def _flow_measured(pl_obj):
        """Return the measured MGD value from the flow report for this limit's averaging period."""
        if not flow_report:
            return None
        ap = (pl_obj.averaging_period or "").lower()
        if ap == "monthly_avg":
            return flow_report.monthly_avg_mgd
        if ap == "daily_max":
            return flow_report.daily_max_mgd
        if ap == "weekly_max":
            return flow_report.weekly_max_mgd
        return flow_report.monthly_avg_mgd  # fallback

    def _flow_limit(pl_obj):
        """Return the permit limit value (MGD) for this flow limit's averaging period."""
        ap = (pl_obj.averaging_period or "").lower()
        if ap == "monthly_avg":
            return pl_obj.monthly_avg_concentration
        if ap == "daily_max":
            return pl_obj.daily_max_concentration
        if ap == "weekly_max":
            return pl_obj.weekly_max_concentration
        return pl_obj.monthly_avg_concentration

    def _sample_violations(pl_obj):
        """Violations that come from sample exceedances (have a sample_id)."""
        viol_rows = (
            db.session.query(Violation, Sample, SampleResult)
            .join(Sample, Violation.sample_id == Sample.id)
            .outerjoin(
                SampleResult,
                (SampleResult.sample_id       == Violation.sample_id) &
                (SampleResult.permit_limit_id == Violation.permit_limit_id),
            )
            .filter(
                Violation.company_id      == company_id,
                Violation.permit_limit_id == pl_obj.id,
                func.extract("month", Violation.violation_date) == month,
                func.extract("year",  Violation.violation_date) == year,
            )
            .order_by(Sample.sample_date)
            .all()
        )
        details = [
            {
                "sample_date":        str(s.sample_date),
                "violation_type":     v.violation_type,
                "exceedance_percent": round(v.exceedance_percent, 2) if v.exceedance_percent is not None else None,
                "severity":           v.violation_severity,
                "concentration":      sr.concentration_result if sr else None,
                "loading":            sr.loading_result       if sr else None,
            }
            for v, s, sr in viol_rows
        ]
        return len(viol_rows), details

    rows = []
    for pl in all_limits:
        param = pl.parameter
        freq  = pl.frequency

        # ── Flow limit: measured value comes from the monthly flow report ──────
        if pl.is_flow_limit:
            measured  = _flow_measured(pl)
            limit_val = _flow_limit(pl)
            exceeds   = (measured is not None and limit_val is not None and measured > limit_val)

            # Flow violations have sample_id=None — query directly
            flow_viols = Violation.query.filter(
                Violation.company_id      == company_id,
                Violation.permit_limit_id == pl.id,
                Violation.violation_type  == "flow_exceeds",
                func.extract("month", Violation.violation_date) == month,
                func.extract("year",  Violation.violation_date) == year,
            ).all()
            n_flow_viols = len(flow_viols)
            flow_viol_details = [
                {
                    "sample_date":        f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}",
                    "violation_type":     v.violation_type,
                    "exceedance_percent": round(v.exceedance_percent, 2) if v.exceedance_percent is not None else None,
                    "severity":           v.violation_severity,
                    "concentration":      measured,
                    "loading":            None,
                }
                for v in flow_viols
            ]

            fr_status = (any_flow_report.review_status if any_flow_report else "missing")

            rows.append({
                "parameter_name":            param.name        if param else "Unknown",
                "abbreviation":              param.abbreviation if param else "",
                "sample_type":               pl.sample_type,
                "frequency":                 freq.description  if freq  else None,
                "is_flow_limit":             True,
                "averaging_period":          pl.averaging_period,
                "flow_report_status":        fr_status,
                "is_monitor_report":         False,
                "is_range_limit":            False,
                "min_value":                 None,
                "max_value":                 None,
                "daily_max_concentration":   pl.daily_max_concentration,
                "daily_max_loading":         None,
                "daily_min_concentration":   None,
                "daily_min_loading":         None,
                "daily_min_is_mr":           False,
                "weekly_max_concentration":  pl.weekly_max_concentration,
                "weekly_max_loading":        None,
                "monthly_avg_concentration": pl.monthly_avg_concentration,
                "monthly_avg_loading":       None,
                "sample_count":              1 if flow_report else 0,
                "exceedance_count":          n_flow_viols,
                "exceedance_details":        flow_viol_details,
                "min_measured":              None,
                "max_measured":              None,
                "avg_measured_conc":         round(measured, 6) if measured is not None else None,
                "avg_measured_load":         None,
                "avg_conc_exceeds":          exceeds,
                "avg_load_exceeds":          False,
            })
            continue

        # ── Standard sample-based limit ────────────────────────────────────────
        entries   = groups.get(pl.id, [])
        n         = len(entries)
        conc_vals = [sr.concentration_result for sr, _ in entries if sr.concentration_result is not None]
        load_vals = [sr.loading_result        for sr, _ in entries if sr.loading_result        is not None]

        avg_conc = (sum(conc_vals) / len(conc_vals)) if conc_vals else None
        avg_load = (sum(load_vals) / len(load_vals)) if load_vals else None

        avg_conc_exceeds = (
            pl.monthly_avg_concentration is not None and
            avg_conc is not None and
            avg_conc > pl.monthly_avg_concentration
        )
        avg_load_exceeds = (
            pl.monthly_avg_loading is not None and
            avg_load is not None and
            avg_load > pl.monthly_avg_loading
        )

        total_exc, exceedance_details = _sample_violations(pl)

        rows.append({
            "parameter_name":            param.name          if param else "Unknown",
            "abbreviation":              param.abbreviation   if param else "",
            "sample_type":               pl.sample_type,
            "frequency":                 freq.description     if freq  else None,
            "is_flow_limit":             False,
            "averaging_period":          None,
            "flow_report_status":        None,
            "is_monitor_report":         pl.is_monitor_report or False,
            "is_range_limit":            pl.is_range_limit    or False,
            "min_value":                 pl.min_value,
            "max_value":                 pl.max_value,
            "daily_max_concentration":   pl.daily_max_concentration,
            "daily_max_loading":         pl.daily_max_loading,
            "daily_min_concentration":   pl.daily_min_concentration,
            "daily_min_loading":         pl.daily_min_loading,
            "daily_min_is_mr":           pl.daily_min_is_mr   or False,
            "weekly_max_concentration":  pl.weekly_max_concentration,
            "weekly_max_loading":        pl.weekly_max_loading,
            "monthly_avg_concentration": pl.monthly_avg_concentration,
            "monthly_avg_loading":       pl.monthly_avg_loading,
            "sample_count":              n,
            "exceedance_count":          total_exc,
            "exceedance_details":        exceedance_details,
            "min_measured":              round(min(conc_vals), 4) if conc_vals else None,
            "max_measured":              round(max(conc_vals), 4) if conc_vals else None,
            "avg_measured_conc":         round(avg_conc, 4)       if avg_conc  is not None else None,
            "avg_measured_load":         round(avg_load, 4)       if avg_load  is not None else None,
            "avg_conc_exceeds":          avg_conc_exceeds,
            "avg_load_exceeds":          avg_load_exceeds,
        })

    permit_number = permits[0].permit_number if permits else None
    return jsonify({
        "company_name":  company.name,
        "permit_number": permit_number,
        "month":         month,
        "year":          year,
        "month_name":    calendar.month_name[month],
        "rows":          rows,
    }), 200


# ── Excel export: sample detail report ────────────────────────────────────────

@reports_bp.route("/samples/export", methods=["GET"])
@login_required
def sample_report_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    company_id   = request.args.get("company_id",   type=int)
    parameter_id = request.args.get("parameter_id", type=int)
    start_date   = request.args.get("start_date")
    end_date     = request.args.get("end_date")

    query = (
        db.session.query(SampleResult, Sample, PermitLimit, Parameter, Company, Permit)
        .join(Sample,      SampleResult.sample_id       == Sample.id)
        .join(PermitLimit, SampleResult.permit_limit_id == PermitLimit.id)
        .join(Parameter,   PermitLimit.parameter_id     == Parameter.id)
        .join(Company,     Sample.company_id             == Company.id)
        .join(Permit,      Sample.permit_id              == Permit.id)
    )
    if company_id:
        query = query.filter(Sample.company_id == company_id)
    if parameter_id:
        query = query.filter(PermitLimit.parameter_id == parameter_id)
    if start_date:
        query = query.filter(Sample.sample_date >= start_date)
    if end_date:
        query = query.filter(Sample.sample_date <= end_date)
    rows = query.order_by(Sample.sample_date, Company.name, Parameter.name).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Detail Report"

    # Styles
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill    = PatternFill("solid", fgColor="1A365D")
    exc_fill    = PatternFill("solid", fgColor="FED7D7")
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="CBD5E0")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "Monitoring Period", "Company", "Permit No.", "Sample Date",
        "Sampler", "Flow (MGD)", "Parameter", "Result (mg/L)", "Loading (lbs/d)",
        "Daily Max (mg/L)", "Daily Max (lbs/d)",
        "Daily Min (mg/L)", "Daily Min (lbs/d)",
        "Weekly Max (mg/L)", "Weekly Max (lbs/d)",
        "Mo. Avg (mg/L)", "Mo. Avg (lbs/d)",
        "Status"
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center
        cell.border    = border

    def _mr_or_val(is_mr, val):
        """Return 'MR' when the MR flag is set, otherwise the numeric value (or None → blank)."""
        return "MR" if is_mr else val

    for sr, s, pl, p, c, permit in rows:
        if pl.is_monitor_report:
            status = "MR"
        elif pl.is_range_limit:
            val = sr.concentration_result
            if val is None:
                status = "—"
            elif pl.min_value is not None and pl.max_value is not None:
                status = "Pass" if pl.min_value <= val <= pl.max_value else "Exceedance"
            else:
                status = "—"
        else:
            exceeds = (
                (pl.daily_max_concentration is not None and sr.concentration_result is not None and sr.concentration_result > pl.daily_max_concentration) or
                (pl.daily_max_loading       is not None and sr.loading_result       is not None and sr.loading_result       > pl.daily_max_loading)
            )
            below_min = not (pl.daily_min_is_mr or False) and (
                (pl.daily_min_concentration is not None and sr.concentration_result is not None and sr.concentration_result < pl.daily_min_concentration) or
                (pl.daily_min_loading       is not None and sr.loading_result       is not None and sr.loading_result       < pl.daily_min_loading)
            )
            status = "Exceedance" if (exceeds or below_min) else "Pass"

        daily_min_mr  = pl.daily_min_is_mr  or False
        weekly_mr     = pl.weekly_max_is_mr or False
        monthly_mr    = pl.monthly_avg_is_mr or False

        ws.append([
            s.sample_date.strftime("%B %Y"),
            c.name,
            permit.permit_number,
            str(s.sample_date),
            s.sampler_name or "",
            round(float(s.flow_mgd), 4) if s.flow_mgd is not None else "",
            p.name,
            sr.concentration_result,
            sr.loading_result,
            pl.daily_max_concentration,
            pl.daily_max_loading,
            _mr_or_val(daily_min_mr, pl.daily_min_concentration),
            _mr_or_val(daily_min_mr, pl.daily_min_loading),
            _mr_or_val(weekly_mr,    pl.weekly_max_concentration),
            _mr_or_val(weekly_mr,    pl.weekly_max_loading),
            _mr_or_val(monthly_mr,   pl.monthly_avg_concentration),
            _mr_or_val(monthly_mr,   pl.monthly_avg_loading),
            status,
        ])
        row_idx = ws.max_row
        is_exc  = status == "Exceedance"
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = border
            if is_exc:
                cell.fill = exc_fill

    col_widths = [16, 22, 14, 13, 18, 11, 20, 14, 14, 16, 16, 16, 16, 16, 16, 16, 16, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "sample_detail_report.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ── PDF export: monthly DMR ────────────────────────────────────────────────────

@reports_bp.route("/monthly/export", methods=["GET"])
@login_required
def monthly_report_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from sqlalchemy import func

    company_id = request.args.get("company_id", type=int)
    month      = request.args.get("month",      type=int)
    year       = request.args.get("year",       type=int)

    if not all([company_id, month, year]):
        return jsonify({"error": "company_id, month, and year are required"}), 400

    company = Company.query.get_or_404(company_id)
    permits = Permit.query.filter_by(company_id=company_id, is_active=True).all()
    permit_ids = [p.id for p in permits]
    permit_number = permits[0].permit_number if permits else "N/A"

    raw = (
        db.session.query(SampleResult, Sample, PermitLimit)
        .join(Sample,      SampleResult.sample_id       == Sample.id)
        .join(PermitLimit, SampleResult.permit_limit_id == PermitLimit.id)
        .filter(
            Sample.company_id == company_id,
            func.extract("month", Sample.sample_date) == month,
            func.extract("year",  Sample.sample_date) == year,
        )
        .all()
    )

    groups: dict = defaultdict(list)
    for sr, s, pl in raw:
        groups[pl.id].append((sr, s))

    all_limits = (
        PermitLimit.query
        .filter(PermitLimit.permit_id.in_(permit_ids))
        .join(Parameter, PermitLimit.parameter_id == Parameter.id)
        .order_by(Parameter.name)
        .all()
    ) if permit_ids else []

    # Reviewed flow report (authoritative) and any flow report (for status messaging)
    flow_report = MonthlyFlowReport.query.filter_by(
        company_id=company_id,
        report_month=month,
        report_year=year,
        review_status="reviewed",
    ).first()
    any_flow_report = flow_report or MonthlyFlowReport.query.filter_by(
        company_id=company_id,
        report_month=month,
        report_year=year,
    ).first()

    def _flow_measured_pdf(pl_obj):
        if not flow_report:
            return None
        ap = (pl_obj.averaging_period or "").lower()
        if ap == "monthly_avg":
            return flow_report.monthly_avg_mgd
        if ap == "daily_max":
            return flow_report.daily_max_mgd
        if ap == "weekly_max":
            return flow_report.weekly_max_mgd
        return flow_report.monthly_avg_mgd

    def _flow_limit_pdf(pl_obj):
        ap = (pl_obj.averaging_period or "").lower()
        if ap == "monthly_avg":
            return pl_obj.monthly_avg_concentration
        if ap == "daily_max":
            return pl_obj.daily_max_concentration
        if ap == "weekly_max":
            return pl_obj.weekly_max_concentration
        return pl_obj.monthly_avg_concentration

    month_name = calendar.month_name[month]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(letter),
                            leftMargin=0.5*inch, rightMargin=0.5*inch,
                            topMargin=0.5*inch, bottomMargin=0.5*inch)

    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle("title", fontSize=14, fontName="Helvetica-Bold", alignment=TA_CENTER, spaceAfter=4)
    sub_s   = ParagraphStyle("sub",   fontSize=10, fontName="Helvetica",      alignment=TA_CENTER, spaceAfter=2)
    cell_s  = ParagraphStyle("cell",  fontSize=7,  fontName="Helvetica",      alignment=TA_LEFT,   leading=9)

    story = [
        Paragraph("Discharge Monitoring Report", title_s),
        Paragraph(f"{company.name}  ·  Permit {permit_number}  ·  {month_name} {year}", sub_s),
        Spacer(1, 10),
    ]

    col_headers = [
        "Parameter", "Freq.", "Samples\nTaken",
        "Daily Max\nLimit (mg/L)", "Daily Max\nLimit (lbs/d)",
        "Daily Min\nLimit (mg/L)",
        "Wkly Max\nLimit (mg/L)",
        "Mo. Avg\nLimit (mg/L)", "Mo. Avg\nLimit (lbs/d)",
        "Min\nMeasured", "Max\nMeasured",
        "Avg\nConc (mg/L)", "Avg\nLoad (lbs/d)",
        "Exceedances", "Status"
    ]

    exc_red = colors.HexColor("#FED7D7")
    hdr_blue = colors.HexColor("#1A365D")

    table_data = [col_headers]
    exc_rows = set()

    def _fmt(v): return str(round(v, 4)) if v is not None else "—"

    for i, pl in enumerate(all_limits, 1):
        param = pl.parameter
        freq  = pl.frequency

        # ── Flow limit: pull from monthly flow report ────────────────────────
        if pl.is_flow_limit:
            measured  = _flow_measured_pdf(pl)
            limit_val = _flow_limit_pdf(pl)
            ap        = (pl.averaging_period or "").lower()
            exceeds   = (measured is not None and limit_val is not None and measured > limit_val)

            n_flow_viols = Violation.query.filter(
                Violation.company_id      == company_id,
                Violation.permit_limit_id == pl.id,
                Violation.violation_type  == "flow_exceeds",
                func.extract("month", Violation.violation_date) == month,
                func.extract("year",  Violation.violation_date) == year,
            ).count()

            is_exc = n_flow_viols > 0 or exceeds
            if is_exc:
                exc_rows.add(i)

            # Limit columns — only show in the column matching averaging period
            d_lim_c  = _fmt(pl.daily_max_concentration)  if ap == "daily_max"   else "—"
            wk_lim_c = _fmt(pl.weekly_max_concentration) if ap == "weekly_max"  else "—"
            mo_lim_c = _fmt(pl.monthly_avg_concentration) if ap == "monthly_avg" else "—"

            fr_status = any_flow_report.review_status if any_flow_report else "missing"
            if fr_status == "reviewed":
                sampled_cell = "Report"
                status_str   = "NON-COMPLIANT" if is_exc else "Compliant"
            elif fr_status == "pending":
                sampled_cell = "Pending"
                status_str   = "Pending"
            else:
                sampled_cell = "No Report"
                status_str   = "No Report"

            table_data.append([
                Paragraph(f"{param.name if param else 'Unknown'} (MGD)", cell_s),
                freq.frequency_code if freq else "—",
                sampled_cell,
                d_lim_c, "—", "—", wk_lim_c,
                mo_lim_c, "—",
                "—", "—",
                _fmt(measured) if measured is not None else "—", "—",
                str(n_flow_viols) if n_flow_viols else "0",
                status_str,
            ])
            continue

        # ── Standard sample-based limit ──────────────────────────────────────
        entries = groups.get(pl.id, [])
        n       = len(entries)
        conc_vals = [sr.concentration_result for sr, _ in entries if sr.concentration_result is not None]
        load_vals = [sr.loading_result        for sr, _ in entries if sr.loading_result        is not None]
        avg_conc  = round(sum(conc_vals)/len(conc_vals), 4) if conc_vals else None
        avg_load  = round(sum(load_vals)/len(load_vals), 4) if load_vals else None

        period_violations = (
            db.session.query(Violation)
            .join(Sample, Violation.sample_id == Sample.id)
            .filter(
                Violation.company_id      == company_id,
                Violation.permit_limit_id == pl.id,
                func.extract("month", Violation.violation_date) == month,
                func.extract("year",  Violation.violation_date) == year,
            ).count()
        )

        avg_exc = (
            (pl.monthly_avg_concentration is not None and avg_conc is not None and avg_conc > pl.monthly_avg_concentration) or
            (pl.monthly_avg_loading       is not None and avg_load is not None and avg_load > pl.monthly_avg_loading)
        )
        total_exc = period_violations + (1 if avg_exc else 0)
        is_exc = total_exc > 0
        if is_exc:
            exc_rows.add(i)

        if pl.is_monitor_report:
            d_lim_c, d_lim_l, d_min_c, wk_max_c, mo_lim_c, mo_lim_l = "MR", "MR", "MR", "MR", "MR", "MR"
        elif pl.is_range_limit:
            mn = pl.min_value if pl.min_value is not None else "—"
            mx = pl.max_value if pl.max_value is not None else "—"
            d_lim_c = f"{mn}–{mx}"
            d_lim_l, d_min_c, wk_max_c, mo_lim_c, mo_lim_l = "—", "—", "—", "—", "—"
        else:
            d_lim_c  = _fmt(pl.daily_max_concentration)
            d_lim_l  = _fmt(pl.daily_max_loading)
            d_min_c  = "MR" if pl.daily_min_is_mr  else _fmt(pl.daily_min_concentration)
            wk_max_c = "MR" if pl.weekly_max_is_mr else _fmt(pl.weekly_max_concentration)
            mo_lim_c = _fmt(pl.monthly_avg_concentration)
            mo_lim_l = _fmt(pl.monthly_avg_loading)

        status = "NON-COMPLIANT" if is_exc else ("No Data" if n == 0 else "Compliant")

        table_data.append([
            Paragraph(param.name if param else "Unknown", cell_s),
            freq.frequency_code if freq else "—",
            str(n),
            d_lim_c, d_lim_l, d_min_c, wk_max_c,
            mo_lim_c, mo_lim_l,
            _fmt(min(conc_vals)) if conc_vals else "—",
            _fmt(max(conc_vals)) if conc_vals else "—",
            _fmt(avg_conc), _fmt(avg_load),
            str(total_exc) if total_exc else "0",
            status,
        ])

    col_widths_pdf = [1.2*inch, 0.42*inch, 0.48*inch,
                      0.70*inch, 0.68*inch, 0.68*inch, 0.68*inch,
                      0.70*inch, 0.68*inch,
                      0.60*inch, 0.60*inch, 0.68*inch, 0.68*inch,
                      0.52*inch, 0.70*inch]

    tbl = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    style_cmds = [
        ("BACKGROUND",   (0,0), (-1,0), hdr_blue),
        ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
        ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",     (0,0), (-1,0), 7),
        ("FONTSIZE",     (0,1), (-1,-1), 7),
        ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ("ALIGN",        (0,1), (0,-1), "LEFT"),
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#F7FAFC")]),
        ("GRID",         (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E0")),
        ("TOPPADDING",   (0,0), (-1,-1), 3),
        ("BOTTOMPADDING",(0,0), (-1,-1), 3),
    ]
    for row_i in exc_rows:
        style_cmds.append(("BACKGROUND", (0, row_i), (-1, row_i), exc_red))
        style_cmds.append(("TEXTCOLOR",  (-1, row_i), (-1, row_i), colors.HexColor("#C53030")))

    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"DMR_{company.name.replace(' ', '_')}_{month_name}_{year}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=filename)
